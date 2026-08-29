"""HRSA Uniform Data System: patients, staffing and payer mix per health center.

Every Section 330 grantee files UDS annually. It is the source for the two
facts a Form 990 never carries, and the two that matter most when sizing an IT
engagement:

* **Patients and visits** -- how big the organization actually is in the only
  unit its own leadership thinks in.
* **Staffing FTEs** -- what drives users, workstations and devices. Revenue is
  a distant proxy for headcount; UDS reports headcount directly.

Plus payer mix (Medicaid, Medicare, uninsured shares), which says something
about funding stability that revenue alone does not.

HRSA publishes this as an annual workbook rather than a stable API, and renames
the columns between years, so this module works the same way as the HRSA site
ingest: a tolerant alias-and-keyword column resolver over whatever file is
present, CSV or XLSX, downloaded by hand into ``uds.local_directory``. Nothing
here is fetched automatically -- the download URLs move, and a file already on
disk is authoritative.

Nothing in UDS is patient-level. Every figure is an organization total.
"""

from __future__ import annotations

import csv
import io
import re
from collections.abc import Callable, Iterable, Iterator, Sequence
from dataclasses import dataclass, field
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import Config
from app.models import (
    IngestRun,
    Organization,
    RunStatus,
    UdsReport,
    utcnow,
)
from pipeline.hrsa import FieldSpec, resolve_columns
from pipeline.text import normalize_name, normalize_state

ProgressFn = Callable[[str], None]


# ---------------------------------------------------------------------------
# Columns
# ---------------------------------------------------------------------------

# UDS renames these between years -- "Total Patients" has also been "Patients
# Served" and "Total Number of Patients" -- so aliases come first and keyword
# groups catch the rest.
UDS_FIELDS: dict[str, FieldSpec] = {
    "hrsa_id": FieldSpec(
        aliases=("BHCMIS Organization Identification Number", "BHCMISID", "Organization ID"),
        contains=(("bhcmis",), ("organization", "identification")),
    ),
    "grant_number": FieldSpec(
        aliases=("Grant Number", "Health Center Grant Number", "BPHC Assigned Number"),
        contains=(("grant", "number"),),
    ),
    "org_name": FieldSpec(
        aliases=("Health Center Name", "Grantee Name", "Organization Name"),
        contains=(("healthcenter", "name"), ("grantee", "name")),
        exclude=("site",),
    ),
    "state": FieldSpec(
        aliases=("State", "State Abbreviation", "Health Center State"),
        contains=(("state",),),
        exclude=("site",),
    ),
    "year": FieldSpec(
        aliases=("Year", "UDS Year", "Reporting Year", "Calendar Year"),
        contains=(("year",),),
    ),
    "patients": FieldSpec(
        aliases=(
            "Total Patients",
            "Patients Served",
            "Total Number of Patients",
            "Total Patients Served",
        ),
        contains=(("total", "patients"), ("patients", "served")),
        exclude=("visit", "percent", "%"),
    ),
    "visits": FieldSpec(
        aliases=("Total Visits", "Total Number of Visits"),
        contains=(("total", "visits"),),
        exclude=("percent", "%"),
    ),
    "sites_reported": FieldSpec(
        aliases=("Total Sites", "Number of Sites", "Service Sites"),
        contains=(("number", "sites"), ("total", "sites")),
    ),
    "total_fte": FieldSpec(
        aliases=("Total FTEs", "Total Personnel FTE", "Total Staff FTE", "Total FTE"),
        contains=(("total", "fte"), ("total", "personnel")),
        exclude=("provider", "physician", "clinical"),
    ),
    "provider_fte": FieldSpec(
        aliases=("Total Provider FTEs", "Provider FTE", "Clinical Staff FTE"),
        contains=(("provider", "fte"), ("clinical", "fte")),
    ),
    "medicaid_share": FieldSpec(
        aliases=("Medicaid Percent", "Percent Medicaid", "Medicaid %"),
        contains=(("medicaid", "percent"), ("medicaid", "%")),
    ),
    "medicare_share": FieldSpec(
        aliases=("Medicare Percent", "Percent Medicare", "Medicare %"),
        contains=(("medicare", "percent"), ("medicare", "%")),
    ),
    "uninsured_share": FieldSpec(
        aliases=("Uninsured Percent", "Percent Uninsured", "Uninsured %"),
        contains=(("uninsured", "percent"), ("uninsured", "%")),
    ),
    "total_revenue": FieldSpec(
        aliases=("Total Revenue", "Total Annual Revenue"),
        contains=(("total", "revenue"),),
        exclude=("grant", "percent", "%"),
    ),
    "grant_revenue": FieldSpec(
        aliases=("BPHC Grant Revenue", "Section 330 Grant Revenue", "Total Grant Revenue"),
        contains=(("grant", "revenue"), ("bphc", "grant")),
    ),
    # HealthCenterInfo in a universal report. The header style there is
    # ProjectDirectorEmail rather than "Project Director Email", which the
    # normalizer already reconciles.
    "director_name": FieldSpec(
        aliases=("Project Director", "ProjectDirector", "Director Name"),
        contains=(("projectdirector",),),
        exclude=("phone", "fax", "email", "ext"),
    ),
    "director_phone": FieldSpec(
        aliases=("Project Director Phone", "ProjectDirectorPhone"),
        contains=(("projectdirector", "phone"),),
        exclude=("ext",),
    ),
    "director_email": FieldSpec(
        aliases=("Project Director Email", "ProjectDirectorEmail"),
        contains=(("projectdirector", "email"),),
    ),
    "urban_rural": FieldSpec(
        aliases=("Urban Rural Flag", "UrbanRuralFlag", "Urban/Rural"),
        contains=(("urban", "rural"),),
    ),
}

REQUIRED_FIELDS = ("patients",)


def missing_required(
    columns: dict[str, str], *, identity_ok: bool = False
) -> list[str]:
    """What a sheet still needs before it is worth reading.

    A file earns its place by carrying patient counts. ``identity_ok`` is the
    single exception, and it is granted only to the HealthCenterInfo sheet of a
    universal report: that sheet has no counts -- they live on the coded table
    sheets -- but it does carry the identity, the address and the project
    director.

    The exception is narrow on purpose. HRSA's site file also has a health
    center name and a BHCMIS id, and allowing identity alone in general made it
    pass as UDS data.
    """
    if "patients" in columns:
        return []
    if identity_ok and "org_name" in columns and (
        "hrsa_id" in columns or "grant_number" in columns
    ):
        return []
    return list(REQUIRED_FIELDS)

# A real UDS workbook is one sheet per UDS table, not a flat export. The
# universal report has 23 of them, and several carry a column that looks like
# "patients" -- Table 6B counts patients *screened* for a condition, which is
# not a patient count at all. So sheets are preferred by name first, and only
# then by whether their columns happen to resolve.
#
# Ranked best-first; matched case-insensitively on a normalized sheet name.
PREFERRED_SHEETS: tuple[str, ...] = (
    "healthcenterinfo",     # identity, address and the project director
    "table3a",              # patients by age and sex
    "table4",               # income and insurance -- payer mix
    "table5",               # staffing and utilization -- FTEs
)

# HealthCenterInfo is the only sheet in a universal report with readable column
# names. The table sheets are UDS form coordinates -- T3a_L1_Ca is Table 3A,
# line 1, column a -- which no amount of keyword matching will decode, so they
# need a form map rather than an alias list. Identity and the project director
# come from here today; the coded tables follow once their line numbers are
# confirmed against real values rather than guessed at.
IDENTITY_SHEET = "healthcenterinfo"
PATIENTS_SHEET = "table3a"

# T3a_L39_Ca is Table 3A, line 39, column a. The line numbers are age bands and
# one of them is the total, but which one has moved between report years -- so
# rather than hard-coding a line number that will quietly become wrong, the
# total is found by arithmetic. See :func:`total_from_coded_lines`.
_CODED_COLUMN = re.compile(
    r"^T(?P<table>[0-9]+[a-z]?)_L(?P<line>[0-9]+[a-z]*)_C(?P<column>[a-z]+)$",
    re.IGNORECASE,
)


def coded_values(row: dict[str, str], table: str) -> dict[str, float]:
    """Values from one coded table in a row, summed across columns per line."""
    per_line: dict[str, float] = {}
    for header, raw in row.items():
        match = _CODED_COLUMN.match(str(header).strip())
        if not match or match.group("table").lower() != table.lower():
            continue
        value = parse_number(raw)
        if value is None:
            continue
        per_line[match.group("line")] = per_line.get(match.group("line"), 0.0) + value
    return per_line


def total_from_coded_lines(per_line: dict[str, float]) -> float | None:
    """The reported total for a coded table, or the sum of its parts.

    A UDS table lists components on numbered lines and repeats their total on
    one more line. Which line that is has changed between report years, and
    guessing it wrong is silent: you get one age band presented as the patient
    count.

    So it is derived instead. If some line equals half the sum of every line,
    that line is the total and the rest are its components -- including the
    total in the sum is what doubles it. With no such line, the table has no
    total row and the sum of the components is the answer.
    """
    if not per_line:
        return None

    grand = sum(per_line.values())
    if grand <= 0:
        return None

    for value in per_line.values():
        # Tolerance for rounding, not for disagreement.
        if abs(value * 2 - grand) <= max(1.0, grand * 0.0005):
            return value
    return grand


def _normalize_sheet(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (name or "").lower())


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


@dataclass
class UdsRecord:
    """One organization-year, before it is matched to an organization."""

    hrsa_id: str | None = None
    grant_number: str | None = None
    name: str | None = None
    state: str | None = None
    year: int | None = None

    patients: int | None = None
    visits: int | None = None
    sites_reported: int | None = None
    total_fte: float | None = None
    provider_fte: float | None = None
    medicaid_share: float | None = None
    medicare_share: float | None = None
    uninsured_share: float | None = None
    total_revenue: float | None = None
    grant_revenue: float | None = None
    director_name: str | None = None
    director_phone: str | None = None
    director_email: str | None = None
    urban_rural: str | None = None

    @property
    def normalized_name(self) -> str:
        return normalize_name(self.name or "")


def parse_number(value: str | None) -> float | None:
    """A number from a spreadsheet cell, or None. Never 0.0 for a blank."""
    text = str(value if value is not None else "").strip()
    if not text or text.lower() in {"n/a", "na", "-", "--", "none", "null", "."}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    cleaned = re.sub(r"[^0-9.\-]", "", text)
    if cleaned in {"", "-", ".", "-."}:
        return None
    try:
        number = float(cleaned)
    except ValueError:
        return None
    return -number if negative else number


def parse_share(value: str | None) -> float | None:
    """A payer-mix share as a 0-1 fraction.

    UDS has published these both as percentages (62.4) and as fractions
    (0.624). A value above 1 is read as a percentage; anything at or below 1 is
    already a fraction. The one genuinely ambiguous case -- exactly 1 -- is read
    as 100%, because a health center with 1% Medicaid is far rarer than one
    with all of it.
    """
    number = parse_number(value)
    if number is None or number < 0:
        return None
    if number > 100:
        return None
    return number / 100.0 if number > 1 else number


def parse_year(value: str | None, fallback: int | None = None) -> int | None:
    number = parse_number(value)
    if number is None:
        return fallback
    year = int(number)
    return year if 1990 <= year <= 2100 else fallback


def year_from_filename(name: str) -> int | None:
    """The reporting year, when the file is named for it and the rows are not.

    UDS exports are routinely called ``2023_UDS_Health_Center_Data.xlsx`` with
    no year column inside, so the filename is the only place the year lives.
    """
    years = [int(m) for m in re.findall(r"(?<!\d)((?:19|20)\d{2})(?!\d)", name)]
    return max(years) if years else None


def read_rows(path: Path) -> tuple[list[str], Iterator[dict[str, str]]]:
    """Header row plus data rows from a CSV, or the best sheet of a workbook."""
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        _name, headers, rows = read_best_sheet(path)
        return headers, rows
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader.fieldnames or []), iter(reader)


def _header_row(values_iter) -> list[str]:
    """The first row that looks like a header rather than a title banner.

    HRSA workbooks routinely open with a merged title cell and a blank line, so
    the header is whichever early row first carries several non-empty cells.
    """
    for values in values_iter:
        filled = [v for v in values if v is not None and str(v).strip()]
        if len(filled) >= 3:
            return [("" if v is None else str(v).strip()) for v in values]
    return []


def sheet_headers(path: Path) -> list[tuple[str, list[str]]]:
    """Every sheet in a workbook, with the header row each one starts with."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [
            (name, _header_row(workbook[name].iter_rows(values_only=True)))
            for name in workbook.sheetnames
        ]
    finally:
        workbook.close()


def read_best_sheet(path: Path) -> tuple[str, list[str], Iterator[dict[str, str]]]:
    """The sheet that actually holds the data, not whichever one was saved last.

    A HRSA workbook opens on a cover sheet -- DataDumpType, ReportingYear, a
    refresh date -- and keeps the health centers on a later one. Reading only
    the active sheet finds four columns of metadata and concludes the file is
    the wrong one, which is exactly backwards.
    """
    candidates = sheet_headers(path)

    def resolves(headers: list[str], name: str = "") -> bool:
        return bool(headers) and not missing_required(
            resolve_columns(headers, UDS_FIELDS),
            identity_ok=_normalize_sheet(name) == IDENTITY_SHEET,
        )

    # A named UDS table wins over whichever sheet happens to resolve first.
    ranked = {name: index for index, name in enumerate(PREFERRED_SHEETS)}
    preferred = sorted(
        (
            (ranked[_normalize_sheet(name)], name)
            for name, _headers in candidates
            if _normalize_sheet(name) in ranked
        ),
    )
    chosen = next(
        (
            name
            for _rank, name in preferred
            if resolves(dict(candidates).get(name, []), name)
        ),
        None,
    )
    if chosen is None:
        chosen = next(
            (name for name, headers in candidates if resolves(headers, name)),
            None,
        )
    if chosen is None:
        # Nothing recognisable: hand back the widest sheet, so the inspector
        # reports the most informative column list it can.
        chosen = max(
            candidates, key=lambda pair: len(pair[1]), default=("", [])
        )[0]

    headers = dict(candidates).get(chosen, [])
    return chosen, headers, _stream_sheet(path, chosen, headers)


def _stream_sheet(
    path: Path, sheet_name: str, headers: list[str]
) -> Iterator[dict[str, str]]:
    """Rows of one sheet as dicts. Opened fresh: a read-only worksheet does not
    reliably survive being iterated twice."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = workbook[sheet_name].iter_rows(values_only=True)
        _header_row(rows)   # advance past the header
        for values in rows:
            if all(v is None or str(v).strip() == "" for v in values):
                continue
            yield {
                headers[i] if i < len(headers) else f"col{i}":
                    ("" if v is None else str(v))
                for i, v in enumerate(values)
            }
    finally:
        workbook.close()


def has_sheet(path: Path, wanted: str) -> bool:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    try:
        return any(
            _normalize_sheet(name) == wanted for name, _headers in sheet_headers(path)
        )
    except Exception:
        return False


@dataclass
class ParseResult:
    records: list[UdsRecord] = field(default_factory=list)
    rows_read: int = 0
    rows_skipped: int = 0
    missing_fields: list[str] = field(default_factory=list)


def parse_uds(
    headers: Sequence[str],
    rows: Iterable[dict[str, str]],
    *,
    default_year: int | None = None,
    identity_ok: bool = False,
) -> ParseResult:
    """Turn one UDS export into records, whatever the year named its columns."""
    columns = resolve_columns(list(headers), UDS_FIELDS)
    result = ParseResult(
        missing_fields=missing_required(columns, identity_ok=identity_ok)
    )
    if result.missing_fields:
        return result

    def cell(row: dict[str, str], key: str) -> str | None:
        column = columns.get(key)
        return row.get(column) if column else None

    for row in rows:
        result.rows_read += 1
        patients = parse_number(cell(row, "patients"))
        name = (cell(row, "org_name") or "").strip()
        hrsa_id = (cell(row, "hrsa_id") or "").strip() or None
        grant_number = (cell(row, "grant_number") or "").strip() or None

        if patients is None and not (hrsa_id or grant_number):
            # Nothing to file it against and nothing to file.
            result.rows_skipped += 1
            continue

        result.records.append(
            UdsRecord(
                hrsa_id=hrsa_id,
                grant_number=grant_number,
                name=name or None,
                state=normalize_state(cell(row, "state") or ""),
                year=parse_year(cell(row, "year"), default_year),
                patients=None if patients is None else int(patients),
                visits=_as_int(parse_number(cell(row, "visits"))),
                sites_reported=_as_int(parse_number(cell(row, "sites_reported"))),
                total_fte=parse_number(cell(row, "total_fte")),
                provider_fte=parse_number(cell(row, "provider_fte")),
                medicaid_share=parse_share(cell(row, "medicaid_share")),
                medicare_share=parse_share(cell(row, "medicare_share")),
                uninsured_share=parse_share(cell(row, "uninsured_share")),
                total_revenue=parse_number(cell(row, "total_revenue")),
                grant_revenue=parse_number(cell(row, "grant_revenue")),
                director_name=(cell(row, "director_name") or "").strip() or None,
                director_phone=(cell(row, "director_phone") or "").strip() or None,
                director_email=(cell(row, "director_email") or "").strip() or None,
                urban_rural=(cell(row, "urban_rural") or "").strip() or None,
            )
        )
    return result


def _as_int(value: float | None) -> int | None:
    return None if value is None else int(value)


def parse_universal(path: Path, *, default_year: int | None = None) -> ParseResult:
    """Read a universal report, whose facts are spread across sheets.

    HealthCenterInfo carries the identity, the address and the project
    director under readable column names. The counts live on coded table
    sheets, joined back on the same identifiers. Everything not yet decoded is
    simply left null rather than approximated from what is.
    """
    sheets = {_normalize_sheet(name): name for name, _h in sheet_headers(path)}

    identity_sheet = sheets[IDENTITY_SHEET]
    _name, headers, rows = (
        identity_sheet,
        dict(sheet_headers(path))[identity_sheet],
        None,
    )
    rows = _stream_sheet(path, identity_sheet, headers)
    result = parse_uds(
        headers, rows, default_year=default_year, identity_ok=True
    )
    if result.missing_fields:
        return result

    by_key: dict[str, UdsRecord] = {}
    for record in result.records:
        key = record.hrsa_id or record.grant_number
        if key:
            by_key[key] = record

    patients_sheet = sheets.get(PATIENTS_SHEET)
    if patients_sheet:
        patient_headers = dict(sheet_headers(path))[patients_sheet]
        for row in _stream_sheet(path, patients_sheet, patient_headers):
            columns = resolve_columns(patient_headers, UDS_FIELDS)
            key = None
            for field_name in ("hrsa_id", "grant_number"):
                column = columns.get(field_name)
                if column and (row.get(column) or "").strip():
                    key = (row.get(column) or "").strip()
                    if key in by_key:
                        break
                    key = None
            record = by_key.get(key) if key else None
            if record is None:
                continue
            total = total_from_coded_lines(coded_values(row, "3a"))
            if total is not None:
                record.patients = int(total)

    return result


# ---------------------------------------------------------------------------
# Matching a UDS row to an organization
# ---------------------------------------------------------------------------


@dataclass
class OrganizationIndex:
    """Lookup tables for joining UDS rows onto the HRSA universe."""

    by_hrsa_id: dict[str, Organization] = field(default_factory=dict)
    by_grant: dict[str, Organization] = field(default_factory=dict)
    by_name_state: dict[tuple[str, str], Organization] = field(default_factory=dict)
    ambiguous_names: set[tuple[str, str]] = field(default_factory=set)

    @classmethod
    def build(cls, organizations: Iterable[Organization]) -> "OrganizationIndex":
        index = cls()
        for organization in organizations:
            if organization.hrsa_id:
                index.by_hrsa_id.setdefault(organization.hrsa_id.strip(), organization)
            if organization.grant_number:
                index.by_grant.setdefault(organization.grant_number.strip(), organization)
            key = (organization.normalized_name, organization.state or "")
            if key in index.by_name_state:
                # Two organizations with the same normalized name in one state
                # cannot be told apart by name, so neither is matched that way.
                index.ambiguous_names.add(key)
            else:
                index.by_name_state[key] = organization
        return index

    def match(self, record: UdsRecord) -> Organization | None:
        """The organization this row belongs to, or None.

        Identifiers first, name last: a name match is a guess and is only used
        when the state agrees and the name is unique within it.
        """
        if record.hrsa_id and record.hrsa_id in self.by_hrsa_id:
            return self.by_hrsa_id[record.hrsa_id]
        if record.grant_number and record.grant_number in self.by_grant:
            return self.by_grant[record.grant_number]
        if record.name and record.state:
            key = (record.normalized_name, record.state)
            if key not in self.ambiguous_names:
                return self.by_name_state.get(key)
        return None


# ---------------------------------------------------------------------------
# Stage
# ---------------------------------------------------------------------------


@dataclass
class UdsResult:
    files_read: int = 0
    rows_read: int = 0
    matched: int = 0
    unmatched: int = 0
    written: int = 0
    messages: list[str] = field(default_factory=list)

    @property
    def status(self) -> RunStatus:
        return RunStatus.SUCCESS


def source_files(directory: Path) -> list[Path]:
    """Every UDS export in the directory, oldest name first."""
    if not directory.exists():
        return []
    return sorted(
        path
        for path in directory.iterdir()
        if path.is_file()
        and not path.name.startswith((".", "~$"))
        and path.suffix.lower() in {".csv", ".xlsx", ".xlsm"}
    )


def ingest(
    session: Session,
    config: Config,
    *,
    on_progress: ProgressFn | None = None,
) -> UdsResult:
    """Load every UDS export in the configured directory."""
    report = on_progress or (lambda _message: None)
    result = UdsResult()

    run = IngestRun(stage="uds", status=RunStatus.RUNNING)
    session.add(run)
    session.commit()

    try:
        directory = config.resolve(config.uds.local_directory)
        files = source_files(directory)
        if not files:
            report(f"No UDS files in {directory}")
            result.messages.append(
                f"No UDS export found in {directory}. Download the health-center "
                "level UDS data from https://data.hrsa.gov/tools/data-reporting "
                "and drop the CSV or XLSX there, then re-run this stage."
            )
            run.status = RunStatus.SUCCESS
            run.finished_at = utcnow()
            run.message = " | ".join(result.messages)
            session.commit()
            return result

        organizations = session.scalars(select(Organization)).all()
        if not organizations:
            result.messages.append(
                "No organizations to attach UDS data to. Run "
                "`python -m pipeline.run --stage hrsa` first."
            )
        index = OrganizationIndex.build(organizations)

        existing = {
            (row.organization_id, row.year): row
            for row in session.scalars(select(UdsReport)).all()
        }

        for path in files:
            default_year = year_from_filename(path.name)
            if has_sheet(path, IDENTITY_SHEET):
                # A universal report: identity on one sheet, counts on others.
                parsed = parse_universal(path, default_year=default_year)
            else:
                headers, rows = read_rows(path)
                parsed = parse_uds(headers, rows, default_year=default_year)
            result.files_read += 1
            result.rows_read += parsed.rows_read

            if parsed.missing_fields:
                result.messages.append(
                    f"{path.name}: no recognizable "
                    f"{', '.join(parsed.missing_fields)} column; skipped"
                )
                report(f"{path.name}: unrecognized layout, skipped")
                continue

            written_here = 0
            for record in parsed.records:
                organization = index.match(record)
                if organization is None:
                    result.unmatched += 1
                    continue
                if record.year is None:
                    # A row with no year cannot be filed against one; the year
                    # is part of the record's identity.
                    result.unmatched += 1
                    continue
                result.matched += 1

                key = (organization.id, record.year)
                row = existing.get(key)
                if row is None:
                    row = UdsReport(organization_id=organization.id, year=record.year)
                    session.add(row)
                    existing[key] = row

                row.patients = record.patients
                row.visits = record.visits
                row.sites_reported = record.sites_reported
                row.total_fte = record.total_fte
                row.provider_fte = record.provider_fte
                row.medicaid_share = record.medicaid_share
                row.medicare_share = record.medicare_share
                row.uninsured_share = record.uninsured_share
                row.total_revenue = record.total_revenue
                row.grant_revenue = record.grant_revenue
                row.director_name = record.director_name
                row.director_phone = record.director_phone
                row.director_email = record.director_email
                row.urban_rural = record.urban_rural
                row.source_file = path.name
                row.fetched_at = utcnow()
                written_here += 1

            result.written += written_here
            report(
                f"{path.name}: {parsed.rows_read:,} rows, "
                f"{written_here:,} attached to organizations"
            )

        session.commit()
    except Exception as exc:
        session.rollback()
        run.status = RunStatus.FAILED
        run.finished_at = utcnow()
        run.message = f"{type(exc).__name__}: {exc}"
        session.commit()
        raise

    if result.unmatched:
        result.messages.append(
            f"{result.unmatched:,} UDS rows did not match an organization in the "
            "HRSA universe (other states, closed grantees, or a missing year column)"
        )

    report(
        f"Stored {result.written:,} organization-years from "
        f"{result.files_read:,} file(s)"
    )

    run.status = result.status
    run.finished_at = utcnow()
    run.records_read = result.rows_read
    run.records_written = result.written
    run.message = " | ".join(result.messages) or None
    session.commit()

    return result


# ---------------------------------------------------------------------------
# Derived sizing
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class SizingEstimate:
    """A starting point for the quantities on a proposal.

    Derived, never reported. The proposal builder prices per Site, User,
    Workstation and Device, and UDS gives the one input that actually predicts
    those numbers: staff headcount. This turns "1,900 patients, 38 FTE" into
    quantities somebody can put in front of a client and argue about -- which is
    the point. It is labelled as an estimate everywhere it appears, and it is
    never written into a field that holds a reported figure.
    """

    sites: int
    users: int
    workstations: int
    devices: int
    basis: str

    def as_dict(self) -> dict[str, object]:
        return {
            "sites": self.sites,
            "users": self.users,
            "workstations": self.workstations,
            "devices": self.devices,
            "basis": self.basis,
        }


def estimate_sizing(
    organization: Organization, report: "UdsReport | None", config: Config
) -> SizingEstimate | None:
    """Suggested proposal quantities, or None when staffing is unknown.

    Returns None rather than guessing from revenue or patient counts: those
    correlate far more weakly with device counts, and a number nobody can
    defend is worse on a proposal than no number at all.
    """
    if report is None or report.total_fte is None or report.total_fte <= 0:
        return None

    fte = float(report.total_fte)
    settings = config.uds
    sites = organization.site_count or report.sites_reported or 1

    return SizingEstimate(
        sites=int(sites),
        # One account per member of staff.
        users=int(round(fte)),
        workstations=max(int(round(fte * settings.workstations_per_fte)), 1),
        devices=max(int(round(fte * settings.devices_per_fte)), 1),
        basis=(
            f"{fte:,.0f} staff FTE reported in the {report.year} UDS, "
            f"at {settings.workstations_per_fte:g} workstations and "
            f"{settings.devices_per_fte:g} devices per FTE"
        ),
    )


# ---------------------------------------------------------------------------
# Inspecting a downloaded file
# ---------------------------------------------------------------------------


def inspect(path: Path) -> str:
    """Describe a downloaded file: is this the right one, and what is in it?

    HRSA publishes many files with similar names and no stable URLs, so the
    fastest way to answer "did I download the right thing" is to point this at
    it and read the answer, rather than running the pipeline and inferring it
    from a row count.
    """
    if not path.exists():
        return f"{path} does not exist."

    if path.name.startswith("~$"):
        # Excel's lock file for a workbook that is currently open. Reporting it
        # as unreadable is technically true and completely unhelpful.
        return (
            f"{path.name}: skipped -- this is Excel's temporary lock file for "
            f"{path.name[2:]}, not a document."
        )

    sheets: list[tuple[str, list[str]]] = []
    used_sheet = ""
    universal = has_sheet(path, IDENTITY_SHEET)
    try:
        if universal:
            sheets = sheet_headers(path)
            used_sheet = dict(
                (_normalize_sheet(name), name) for name, _h in sheets
            )[IDENTITY_SHEET]
            headers = dict(sheets)[used_sheet]
            parsed = parse_universal(
                path, default_year=year_from_filename(path.name)
            )
        elif path.suffix.lower() in {".xlsx", ".xlsm"}:
            sheets = sheet_headers(path)
            used_sheet, headers, rows = read_best_sheet(path)
            parsed = parse_uds(headers, rows, default_year=year_from_filename(path.name))
        else:
            headers, rows = read_rows(path)
            parsed = parse_uds(headers, rows, default_year=year_from_filename(path.name))
    except Exception as exc:
        return f"{path.name}: could not be read as CSV or Excel ({exc})"
    lines = [f"{path.name}  ({path.stat().st_size / 1_048_576:.1f} MB)"]

    if len(sheets) > 1:
        lines.append("")
        note = (
            " (a universal report: identity here, counts joined from the "
            "coded table sheets)"
            if universal
            else ""
        )
        lines.append(f"  {len(sheets)} sheets; reading \"{used_sheet}\"{note}:")
        for name, sheet_columns in sheets:
            mark = ">" if name == used_sheet else " "
            lines.append(f"    {mark} {name}  ({len(sheet_columns)} columns)")

    if parsed.missing_fields:
        lines.append("")
        lines.append("  NOT a UDS health-center file.")
        lines.append(
            f"  No column that looks like: {', '.join(parsed.missing_fields)}."
        )
        lines.append("")
        # Every column, not a sample. This is the moment somebody needs the
        # full list -- either to recognise the file, or to send it on so the
        # aliases can be taught the names this year's export uses.
        lines.append(f"  All {len(headers)} columns:")
        for index, header in enumerate(headers, start=1):
            lines.append(f"    {index:3d}. {header}")
        lines.append("")
        lines.append(
            "  A usable file has one row per health center and a total-patients "
            "column. If this is a national or state summary, it is the wrong one."
        )
        return "\n".join(lines)

    found = resolve_columns(list(headers), UDS_FIELDS)
    have = [key for key in UDS_FIELDS if key in found]
    if universal and any(r.patients is not None for r in parsed.records):
        have.append("patients (from Table3A)")
    missing = [
        key
        for key in UDS_FIELDS
        if key not in found
        and not (key == "patients" and "patients (from Table3A)" in have)
    ]

    lines.append("")
    lines.append(f"  Looks like UDS data: {parsed.rows_read:,} rows.")
    years = sorted({r.year for r in parsed.records if r.year})
    lines.append(
        f"  Reporting year: {', '.join(str(y) for y in years)}"
        if years
        else "  Reporting year: not in the file -- rename it to include the year, "
        "e.g. 2023_UDS.csv"
    )
    lines.append(f"  Will read: {', '.join(have)}")
    if missing:
        lines.append(f"  Not present (left blank): {', '.join(missing)}")

    sample = [r for r in parsed.records if r.name][:3]
    if sample:
        lines.append("")
        lines.append("  First few organizations:")
        for record in sample:
            patients = f"{record.patients:,}" if record.patients is not None else "?"
            lines.append(
                f"    {record.name} ({record.state or '??'}) -- {patients} patients"
            )
    return "\n".join(lines)


def preview_sheet(path: Path, sheet: str, rows: int = 2) -> str:
    """Column names beside the values in the first rows.

    A coded sheet's headers say nothing -- T3a_L39_Ca could be any of thirty-nine
    age bands or the total. The values tell you which, and a total is obvious on
    sight next to its components.
    """
    try:
        sheets = dict(sheet_headers(path))
    except Exception as exc:
        return f"{path.name}: could not be read ({exc})"

    by_normalized = {_normalize_sheet(name): name for name in sheets}
    actual = by_normalized.get(_normalize_sheet(sheet))
    if actual is None:
        return f"No sheet called {sheet!r}. Available: {', '.join(sheets)}"

    headers = sheets[actual]
    lines = [f"{path.name} -- {actual}"]
    for index, row in enumerate(_stream_sheet(path, actual, headers)):
        if index >= rows:
            break
        lines.append("")
        lines.append(f"  Row {index + 1}:")
        for column, value in row.items():
            text = str(value).strip()
            lines.append(f"    {column:<28} {text[:60]}")
    return "\n".join(lines)


def describe_sheets(path: Path, wanted: Sequence[str]) -> str:
    """Every column of the named sheets, numbered.

    The way to teach the reader a layout it has not seen: run this, send the
    output, and the column names become aliases.
    """
    try:
        sheets = dict(sheet_headers(path))
    except Exception as exc:
        return f"{path.name}: could not be read ({exc})"

    lines = [f"{path.name}"]
    by_normalized = {_normalize_sheet(name): name for name in sheets}

    for request in wanted:
        actual = by_normalized.get(_normalize_sheet(request))
        if actual is None:
            lines.append("")
            lines.append(f"  No sheet called {request!r}. Available:")
            lines.extend(f"    {name}" for name in sheets)
            continue
        headers = sheets[actual]
        lines.append("")
        lines.append(f"  {actual}  ({len(headers)} columns)")
        for index, header in enumerate(headers, start=1):
            lines.append(f"    {index:4d}. {header}")
    return "\n".join(lines)


def main(argv: list[str] | None = None) -> int:
    import argparse

    parser = argparse.ArgumentParser(
        prog="python -m pipeline.uds",
        description="Check whether a downloaded file is usable UDS data.",
    )
    parser.add_argument("path", type=Path, nargs="+", help="File(s) to inspect.")
    parser.add_argument(
        "--sheet",
        action="append",
        default=[],
        metavar="NAME",
        help=(
            "Print every column of this sheet instead of inspecting the file. "
            "Repeatable. Use it to show what a workbook actually contains."
        ),
    )
    parser.add_argument(
        "--preview",
        metavar="NAME",
        help=(
            "Print the first rows of this sheet with each value beside its "
            "column name. Use it on a sheet whose headers are form codes."
        ),
    )
    parser.add_argument(
        "--rows", type=int, default=2, help="Rows to show with --preview."
    )
    args = parser.parse_args(argv)

    for path in args.path:
        if args.preview:
            print(preview_sheet(path, args.preview, rows=args.rows))
        elif args.sheet:
            print(describe_sheets(path, args.sheet))
        else:
            print(inspect(path))
        print()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
