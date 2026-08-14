"""End-to-end acceptance: every stage wired together, then the UI on top.

The other test modules each cover one module in isolation. This one runs the
real CLI stage list -- HRSA, EIN matching, financials, scoring, changes -- from
an empty database against fixture data, then boots the web app on the result and
checks the pages a user actually looks at. It is the test that catches wiring
mistakes between modules rather than inside them.
"""

from __future__ import annotations

import io
import json
import shutil
from pathlib import Path

import httpx
import pytest
import yaml
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.config import Config, load_config
from app.db import init_db, reset_engine, session_scope
from app.models import (
    ChangeEvent,
    ChangeKind,
    EinMatch,
    Filing,
    MatchStatus,
    Organization,
    Score,
)
from pipeline.run import STAGES, StageOptions

FIXTURES = Path(__file__).parent / "fixtures"
ORG_PAYLOAD = json.loads((FIXTURES / "propublica_organization.json").read_text())

# Search results keyed by a fragment of the query, mirroring the shape the live
# endpoint returns.
SEARCH_RESULTS = {
    "erie": [
        {
            "ein": 362167869,
            "name": "ERIE FAMILY HEALTH CENTERS INC",
            "city": "Chicago",
            "state": "IL",
            "have_filings": True,
        }
    ],
    "milwaukee": [
        {
            "ein": 391385403,
            "name": "MILWAUKEE HEALTH SERVICES INC",
            "city": "Milwaukee",
            "state": "WI",
            "have_filings": True,
        }
    ],
    "prairie": [
        {
            "ein": 351122334,
            "name": "PRAIRIE RURAL HEALTH SERVICES",
            "city": "Lafayette",
            "state": "IN",
        }
    ],
}


def scaled_payload(ein: str, scale: float) -> dict:
    payload = json.loads(json.dumps(ORG_PAYLOAD))
    for filing in payload["filings_with_data"]:
        for key in ("totrevenue", "totfuncexpns", "totassetsend"):
            filing[key] = round(filing[key] * scale)
    return payload


SCALES = {"362167869": 0.25, "391385403": 0.12, "351122334": 0.02}


def api_handler(request: httpx.Request) -> httpx.Response:
    url = str(request.url)
    if "/search.json" in url:
        query = request.url.params.get("q", "").lower()
        for keyword, organizations in SEARCH_RESULTS.items():
            if keyword in query:
                return httpx.Response(
                    200,
                    json={
                        "total_results": len(organizations),
                        "organizations": organizations,
                    },
                )
        return httpx.Response(200, json={"total_results": 0, "organizations": []})

    ein = url.rsplit("/", 1)[-1].replace(".json", "").zfill(9)
    if ein not in SCALES:
        return httpx.Response(404)
    return httpx.Response(200, json=scaled_payload(ein, SCALES[ein]))


@pytest.fixture()
def project(tmp_path: Path, config: Config, monkeypatch):
    """An isolated config, cache and database, with the API mocked."""
    raw = yaml.safe_load((config.project_root / "config.yaml").read_text())
    raw["app"]["database_path"] = str(tmp_path / "e2e.db")
    raw["cache"]["directory"] = str(tmp_path / "raw")
    config_path = tmp_path / "config.yaml"
    config_path.write_text(yaml.safe_dump(raw))

    cache = tmp_path / "raw"
    cache.mkdir(parents=True)
    shutil.copy(FIXTURES / "hrsa_sites.csv", cache / "hrsa_service_delivery_sites.csv")
    shutil.copy(FIXTURES / "hrsa_awardees.csv", cache / "hrsa_program_awardees.csv")

    loaded = load_config(config_path)

    # Both API stages build their own client; hand them a mocked transport.
    import pipeline.run as run_module
    from pipeline.propublica import ProPublicaClient

    def make_client(cfg, session, **_kwargs):
        return ProPublicaClient(
            cfg,
            session,
            client=httpx.Client(transport=httpx.MockTransport(api_handler)),
            sleep=lambda _s: None,
            monotonic=lambda: 0.0,
        )

    monkeypatch.setattr(run_module, "ProPublicaClient", make_client)

    reset_engine()
    init_db(loaded)
    yield loaded, cache
    reset_engine()


def run_pipeline(config: Config, *, stages: tuple[str, ...] | None = None) -> list[str]:
    """Run the real stage list and return the progress messages."""
    messages: list[str] = []
    selected = [s for s in STAGES if not stages or s.name in stages]
    with session_scope(config) as session:
        for stage in selected:
            stage.run(session, config, StageOptions(), messages.append)
    return messages


# ---------------------------------------------------------------------------
# A complete first run
# ---------------------------------------------------------------------------


def test_full_pipeline_from_empty_to_scored(project) -> None:
    config, _ = project
    run_pipeline(config)

    with session_scope(config) as session:
        organizations = session.scalars(select(Organization)).all()
        assert len(organizations) == 5

        # Every organization is scored, including those with no financials.
        assert session.scalar(select(func.count()).select_from(Score)) == 5

        erie = next(o for o in organizations if o.name.startswith("Erie"))
        assert erie.site_count == 3
        assert erie.federal_award_amount == 6_200_000.0
        assert erie.ein == "362167869"

        filings = session.scalars(
            select(Filing).where(Filing.ein == "362167869").order_by(Filing.tax_year.desc())
        ).all()
        assert [f.tax_year for f in filings] == [2023, 2022, 2021]
        assert filings[0].total_revenue == pytest.approx(89_412_355 * 0.25)

        # The first changes run is a baseline, so nothing is reported yet.
        assert session.scalar(select(func.count()).select_from(ChangeEvent)) == 0


def test_out_of_footprint_organizations_are_scored_but_not_searched(project) -> None:
    """All five fixture organizations sit in the footprint, so add one that does not."""
    config, _ = project
    with session_scope(config) as session:
        session.add(
            Organization(
                dedup_key="lone star health|TX",
                name="Lone Star Health",
                normalized_name="lone star health",
                state="TX",
                city="Austin",
                site_count=6,
            )
        )

    run_pipeline(config)

    with session_scope(config) as session:
        texas = session.scalars(
            select(Organization).where(Organization.state == "TX")
        ).one()
        # Scored on what is known locally, but never sent to the API.
        assert texas.score is not None
        assert texas.ein_match is None


def test_unconfirmed_matches_carry_no_financials_end_to_end(project) -> None:
    """The central rule, checked through the whole stack rather than a unit."""
    config, _ = project
    run_pipeline(config)

    with session_scope(config) as session:
        pending = session.scalars(
            select(Organization)
            .join(EinMatch)
            .where(EinMatch.status == MatchStatus.PENDING.value)
        ).all()
        for organization in pending:
            assert organization.ein is None
            score = organization.score
            assert "revenue" in (score.unavailable_factors or [])


# ---------------------------------------------------------------------------
# A second run a month later
# ---------------------------------------------------------------------------


def test_second_run_detects_real_movement(project) -> None:
    config, cache = project
    run_pipeline(config)

    # HRSA publishes a new release: Erie has opened a fourth site.
    sites = (cache / "hrsa_service_delivery_sites.csv").read_text()
    new_row = (
        '010010,"Erie Family Health Centers, Inc.",Erie Pilsen Health Center,S-1005,'
        "1836 W 18th St,Chicago,IL,60608,(312) 666-3494,,Service Delivery Site,Active,"
        "Health Center Program Awardee,H80CS00123\n"
    )
    (cache / "hrsa_service_delivery_sites.csv").write_text(sites + new_row)

    run_pipeline(config)

    with session_scope(config) as session:
        erie = session.scalars(
            select(Organization).where(Organization.name.startswith("Erie"))
        ).one()
        assert erie.site_count == 4

        events = session.scalars(select(ChangeEvent)).all()
        assert len(events) == 1
        assert events[0].kind == ChangeKind.SITES
        assert events[0].summary == "Opened 1 delivery site (3 to 4)"
        assert events[0].organization_id == erie.id


def test_rerunning_unchanged_data_reports_nothing(project) -> None:
    config, _ = project
    run_pipeline(config)
    run_pipeline(config)

    with session_scope(config) as session:
        assert session.scalar(select(func.count()).select_from(ChangeEvent)) == 0


def test_human_decisions_survive_a_full_rerun(project) -> None:
    """A rejected match must not be quietly re-proposed by the next run."""
    config, _ = project
    run_pipeline(config)

    with session_scope(config) as session:
        pending = session.scalars(
            select(EinMatch).where(EinMatch.status == MatchStatus.PENDING.value)
        ).first()
        assert pending is not None
        organization_id = pending.organization_id
        pending.rejected_eins = [pending.ein]
        pending.ein = None
        pending.status = MatchStatus.REJECTED
        pending.decided_by = "analyst"

    run_pipeline(config)

    with session_scope(config) as session:
        match = session.scalars(
            select(EinMatch).where(EinMatch.organization_id == organization_id)
        ).one()
        assert match.decided_by == "analyst"
        assert match.status == MatchStatus.REJECTED


# ---------------------------------------------------------------------------
# The UI on top of a real pipeline result
# ---------------------------------------------------------------------------


@pytest.fixture()
def client(project):
    config, cache = project
    run_pipeline(config)

    from app import main
    from app.db import get_session_factory

    factory = get_session_factory(config)
    main.app.dependency_overrides[main.get_db] = lambda: factory()
    with TestClient(main.app) as test_client:
        yield test_client, config, cache
    main.app.dependency_overrides.clear()


def test_dashboard_reflects_the_pipeline_result(client) -> None:
    test_client, _, _ = client
    body = test_client.get("/").text

    assert "Erie Family Health Centers, Inc." in body
    assert "Top 10 prospects" in body
    # Branded furniture survives the whole stack.
    assert body.count("star-band__block--") == 6
    assert "954 W. Washington Blvd" in body


def test_every_organization_page_renders(client) -> None:
    test_client, config, _ = client
    with session_scope(config) as session:
        ids = [o.id for o in session.scalars(select(Organization)).all()]

    for organization_id in ids:
        response = test_client.get(f"/organizations/{organization_id}")
        assert response.status_code == 200, organization_id
        # Nothing renders a raw None or an unrendered template expression.
        assert "None" not in response.text.replace("None available", "")
        assert "{{" not in response.text


def test_review_queue_and_a_decision_flow_through(client) -> None:
    test_client, config, _ = client

    with session_scope(config) as session:
        pending = session.scalars(
            select(EinMatch).where(EinMatch.status == MatchStatus.PENDING.value)
        ).first()
        assert pending is not None
        organization_id, ein = pending.organization_id, pending.ein

    assert "EIN review queue" in test_client.get("/review").text

    response = test_client.post(
        f"/review/{organization_id}/accept",
        data={"ein": ein, "decided_by": "analyst"},
        follow_redirects=False,
    )
    assert response.status_code == 303

    with session_scope(config) as session:
        organization = session.get(Organization, organization_id)
        assert organization.ein == ein
        assert organization.ein_match.status == MatchStatus.ACCEPTED


def test_accepting_a_match_lets_the_next_run_enrich_it(client) -> None:
    """Confirming an EIN must actually pull that organization's filings."""
    test_client, config, _ = client

    with session_scope(config) as session:
        pending = session.scalars(
            select(EinMatch).where(EinMatch.status == MatchStatus.PENDING.value)
        ).first()
        organization_id, ein = pending.organization_id, pending.ein
        assert not session.scalars(select(Filing).where(Filing.ein == ein)).all()

    test_client.post(
        f"/review/{organization_id}/accept", data={"ein": ein}, follow_redirects=False
    )
    run_pipeline(config, stages=("financials", "scoring"))

    with session_scope(config) as session:
        assert session.scalars(select(Filing).where(Filing.ein == ein)).all()
        organization = session.get(Organization, organization_id)
        assert "revenue" not in (organization.score.unavailable_factors or [])


def test_exports_carry_the_pipeline_result(client) -> None:
    test_client, _, _ = client

    csv_body = test_client.get("/export.csv?state=IL").text
    assert csv_body.splitlines()[0] == "Allstar Partners"
    assert "Erie Family Health Centers, Inc." in csv_body
    assert "Milwaukee Health Services" not in csv_body

    workbook = load_workbook(io.BytesIO(test_client.get("/export.xlsx").content))
    sheet = workbook["Prospects"]
    names = [row[1] for row in sheet.iter_rows(min_row=1, values_only=True)]
    assert any(name and "Erie" in str(name) for name in names)


def test_changes_page_renders_after_a_second_run(client) -> None:
    test_client, config, cache = client

    sites = (cache / "hrsa_service_delivery_sites.csv").read_text()
    (cache / "hrsa_service_delivery_sites.csv").write_text(
        sites
        + '010010,"Erie Family Health Centers, Inc.",Erie Pilsen Health Center,S-1005,'
        "1836 W 18th St,Chicago,IL,60608,,,Service Delivery Site,Active,"
        "Health Center Program Awardee,H80CS00123\n"
    )
    run_pipeline(config, stages=("hrsa", "scoring", "changes"))

    body = test_client.get("/changes").text
    assert "Opened 1 delivery site" in body
    assert "Erie Family Health Centers, Inc." in body
