"""Dashboard routes, filtering, review decisions and exports."""

from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, sessionmaker

from app.config import Config
from app.db import create_db_engine
from app.models import (
    Base,
    EinMatch,
    Filing,
    GranteeType,
    IngestRun,
    MatchStatus,
    Organization,
    Person,
    RunStatus,
    Score,
    WebsitePerson,
    utcnow,
)
from app.queries import Filters, data_status, fetch_rows, review_queue, summarize


@pytest.fixture()
def populated(tmp_path: Path, config: Config):
    """A small but representative database, plus a client bound to it."""
    engine = create_db_engine(f"sqlite:///{tmp_path / 'web.db'}")
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, expire_on_commit=False, future=True)
    session = factory()

    def add(
        name,
        state,
        city,
        sites,
        award,
        ein,
        status,
        composite,
        revenue=None,
        match_score=95.0,
        grantee_type=GranteeType.AWARDEE,
    ):
        org = Organization(
            dedup_key=f"{name.lower()}|{state}",
            name=name,
            normalized_name=name.lower(),
            state=state,
            city=city,
            site_count=sites,
            federal_award_amount=award,
            grantee_type=grantee_type,
        )
        session.add(org)
        session.flush()
        if ein or status is not None:
            session.add(
                EinMatch(
                    organization_id=org.id,
                    ein=ein,
                    score=match_score,
                    status=status,
                    matched_name=f"{name.upper()} INC",
                    matched_city=city,
                    matched_state=state,
                    candidates=[
                        {
                            "ein": ein or "999999999",
                            "name": f"{name.upper()} INC",
                            "city": city,
                            "state": state,
                            "score": match_score,
                            "city_match": True,
                            "has_filings": True,
                        }
                    ],
                )
            )
        if composite is not None:
            session.add(
                Score(
                    organization_id=org.id,
                    composite=composite,
                    breakdown=[
                        {
                            "factor": "revenue",
                            "label": "Annual revenue",
                            "score": None if revenue is None else 100,
                            "weight": 35,
                            "effective_weight": 0 if revenue is None else 35,
                            "available": revenue is not None,
                            "detail": "test",
                            "value": "test",
                        }
                    ],
                    unavailable_factors=[] if revenue is not None else ["revenue"],
                )
            )
        if revenue is not None and ein:
            session.add(
                Filing(
                    ein=ein,
                    tax_year=2023,
                    total_revenue=revenue,
                    total_expenses=revenue * 0.95,
                    total_assets=revenue * 0.8,
                    pdf_url="https://example.org/990.pdf",
                    period_end=datetime(2023, 12, 31, tzinfo=timezone.utc),
                    form_type="990",
                )
            )
        return org

    erie = add("Erie Family Health", "IL", "Chicago", 12, 6_000_000, "362167869", MatchStatus.AUTO, 88.0, 20_000_000)
    # Erie has people from its filing; Milwaukee only from its website. Both
    # shapes need to survive into the contacts export.
    session.add(
        Person(
            ein="362167869",
            tax_year=2023,
            name="MARIA T ALVAREZ",
            title="CHIEF EXECUTIVE OFFICER",
            roles=["Officer"],
            compensation=451_000,
        )
    )
    milwaukee = add("Milwaukee Health Services", "WI", "Milwaukee", 4, None, "391385403", MatchStatus.ACCEPTED, 71.0, 8_000_000, grantee_type=GranteeType.LOOK_ALIKE)
    session.add(
        WebsitePerson(
            organization_id=milwaukee.id,
            name="Denise Whitaker",
            title="Board Chair",
            email="dwhitaker@example.org",
            source_url="https://example.org/board",
        )
    )
    add("Prairie Rural Health", "IN", "Lafayette", 1, None, "351122334", MatchStatus.PENDING, 44.0, None, match_score=77.0)
    add("Riverbend Access", "IL", "Peoria", 2, None, None, MatchStatus.UNMATCHED, 40.0, None, match_score=None)
    add("Northwoods Clinic", "MI", "Marquette", 6, None, None, MatchStatus.PENDING, 55.0, None, match_score=82.0)

    session.add(
        IngestRun(
            stage="hrsa",
            status=RunStatus.SUCCESS,
            started_at=utcnow() - timedelta(minutes=5),
            finished_at=utcnow(),
            records_written=5,
        )
    )
    session.commit()

    from app import main

    main.app.dependency_overrides[main.get_db] = lambda: factory()
    client = TestClient(main.app)
    try:
        yield client, session, erie
    finally:
        main.app.dependency_overrides.clear()
        session.close()
        engine.dispose()


# ---------------------------------------------------------------------------
# Pages render
# ---------------------------------------------------------------------------


def test_dashboard_renders_with_brand_furniture(populated) -> None:
    client, _, _ = populated
    response = client.get("/")
    body = response.text

    assert response.status_code == 200
    # Star band, footer, brand stylesheet -- the guide's required elements.
    assert body.count("star-band__block--") == 6
    assert "954 W. Washington Blvd. Ste 535, Chicago, IL 60607" in body
    assert "Information Technology Consultants" in body
    assert "css/brand.css" in body


def test_dashboard_shows_the_summary_strip(populated) -> None:
    client, _, _ = populated
    body = client.get("/").text

    assert "FQHC organizations" in body
    assert "EIN matched" in body
    assert "Average ICP score" in body
    assert "Needs review" in body
    assert "Top 10 prospects" in body


def test_detail_page_shows_financials_and_freshness(populated) -> None:
    client, _, erie = populated
    body = client.get(f"/organizations/{erie.id}").text

    assert "Erie Family Health" in body
    assert "FY2023" in body
    assert "$20.0M" in body
    assert "990 PDF" in body
    assert "36-2167869" in body


def test_unknown_organization_returns_404(populated) -> None:
    client, _, _ = populated
    assert client.get("/organizations/99999").status_code == 404


def test_missing_data_is_labelled_not_available(populated) -> None:
    """The rule that matters most, checked at the rendering layer."""
    client, session, _ = populated
    prairie = session.scalars(
        select(Organization).where(Organization.name == "Prairie Rural Health")
    ).one()

    body = client.get(f"/organizations/{prairie.id}").text
    assert "Not available" in body
    # An unconfirmed match must explain why there are no financials.
    assert "no confirmed EIN" in body


def test_review_page_lists_pending_matches(populated) -> None:
    client, _, _ = populated
    body = client.get("/review").text

    assert "EIN review queue" in body
    assert "Prairie Rural Health" in body
    assert "Northwoods Clinic" in body
    assert "Erie Family Health" not in body.split("EIN review queue")[1]


# ---------------------------------------------------------------------------
# Filtering and sorting
# ---------------------------------------------------------------------------


def test_state_filter_narrows_the_table(populated) -> None:
    client, _, _ = populated
    body = client.get("/table?state=WI").text

    assert "Milwaukee Health Services" in body
    assert "Erie Family Health" not in body


def test_search_matches_name_and_city(populated) -> None:
    client, _, _ = populated
    assert "Erie Family Health" in client.get("/table?q=erie").text
    assert "Milwaukee Health Services" in client.get("/table?q=milwaukee").text
    assert "Erie Family Health" not in client.get("/table?q=milwaukee").text


def test_min_score_filter(populated) -> None:
    client, _, _ = populated
    body = client.get("/table?min_score=70").text

    assert "Erie Family Health" in body
    assert "Riverbend Access" not in body


def test_revenue_filter_only_matches_organizations_with_known_revenue(
    populated,
) -> None:
    """An unknown revenue must not sneak into a revenue-bounded filter."""
    client, _, _ = populated
    body = client.get("/table?min_revenue=1000000").text

    assert "Erie Family Health" in body
    assert "Prairie Rural Health" not in body


def test_match_status_filter(populated) -> None:
    client, _, _ = populated

    confirmed = client.get("/table?match=confirmed").text
    assert "Erie Family Health" in confirmed
    assert "Prairie Rural Health" not in confirmed

    review = client.get("/table?match=review").text
    assert "Prairie Rural Health" in review
    assert "Erie Family Health" not in review


def test_sorting_by_revenue_puts_unknown_last(populated, session: Session) -> None:
    """Nulls last in both directions: unknown is not the same as smallest."""
    _, db, _ = populated

    ascending, _ = fetch_rows(db, Filters(sort="revenue", direction="asc"))
    descending, _ = fetch_rows(db, Filters(sort="revenue", direction="desc"))

    assert [r.revenue for r in ascending][:2] == [8_000_000, 20_000_000]
    assert [r.revenue for r in descending][:2] == [20_000_000, 8_000_000]
    assert ascending[-1].revenue is None
    assert descending[-1].revenue is None


def test_pending_matches_expose_no_revenue(populated) -> None:
    """Financials must not reach the table through an unconfirmed EIN."""
    _, db, _ = populated
    db.add(Filing(ein="351122334", tax_year=2023, total_revenue=99_000_000))
    db.commit()

    rows, _ = fetch_rows(db, Filters())
    prairie = next(r for r in rows if r.organization.name == "Prairie Rural Health")
    assert prairie.revenue is None


def test_summary_counts(populated) -> None:
    _, db, _ = populated
    summary = summarize(db)

    assert summary.total == 5
    assert summary.matched == 2
    assert summary.needs_review == 2
    assert summary.with_financials == 2
    assert summary.total_sites == 25


def test_review_queue_orders_least_certain_first(populated) -> None:
    _, db, _ = populated
    queue = review_queue(db)
    assert [row.match.score for row in queue] == [77.0, 82.0]


# ---------------------------------------------------------------------------
# Review decisions
# ---------------------------------------------------------------------------


def test_accepting_a_match_confirms_the_ein(populated) -> None:
    client, db, _ = populated
    prairie = db.scalars(
        select(Organization).where(Organization.name == "Prairie Rural Health")
    ).one()

    response = client.post(
        f"/review/{prairie.id}/accept",
        data={"ein": "351122334", "decided_by": "analyst"},
        follow_redirects=False,
    )
    assert response.status_code == 303

    db.expire_all()
    prairie = db.get(Organization, prairie.id)
    assert prairie.ein_match.status == MatchStatus.ACCEPTED
    assert prairie.ein == "351122334"      # now usable for financials
    assert prairie.ein_match.decided_by == "analyst"
    assert prairie.ein_match.decided_at is not None


def test_rejecting_a_match_records_the_ein_and_clears_it(populated) -> None:
    client, db, _ = populated
    prairie = db.scalars(
        select(Organization).where(Organization.name == "Prairie Rural Health")
    ).one()

    client.post(
        f"/review/{prairie.id}/reject",
        data={"ein": "351122334"},
        follow_redirects=False,
    )

    db.expire_all()
    prairie = db.get(Organization, prairie.id)
    assert prairie.ein_match.status == MatchStatus.REJECTED
    assert prairie.ein_match.ein is None
    assert prairie.ein is None
    # Recorded so the next pipeline run proposes a different candidate.
    assert prairie.ein_match.rejected_eins == ["351122334"]


def test_fetch_requests_get_a_fragment_instead_of_a_redirect(populated) -> None:
    client, db, _ = populated
    prairie = db.scalars(
        select(Organization).where(Organization.name == "Prairie Rural Health")
    ).one()

    response = client.post(
        f"/review/{prairie.id}/accept",
        data={"ein": "351122334"},
        headers={"X-Requested-With": "fetch"},
    )
    assert response.status_code == 200
    assert "Match confirmed" in response.text


def test_decision_on_a_missing_organization_is_404(populated) -> None:
    client, _, _ = populated
    response = client.post(
        "/review/99999/accept", data={"ein": "111111111"}, follow_redirects=False
    )
    assert response.status_code == 404


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------


def test_csv_export_has_the_branded_header_block(populated) -> None:
    client, _, _ = populated
    response = client.get("/export.csv")
    body = response.text

    assert response.status_code == 200
    assert "attachment;" in response.headers["content-disposition"]
    lines = body.splitlines()
    assert lines[0] == "Allstar Partners"
    assert lines[1] == "FQHC Prospect Intelligence"
    assert lines[2].startswith("Generated ")
    assert lines[3].startswith("Filters: none")
    assert "ICP score,Organization" in body


def test_csv_export_respects_the_current_filter(populated) -> None:
    client, _, _ = populated
    body = client.get("/export.csv?state=WI").text

    assert "Milwaukee Health Services" in body
    assert "Erie Family Health" not in body
    assert "state in WI" in body  # the filter is stated in the file


def test_csv_export_writes_not_available_for_gaps(populated) -> None:
    """A blank cell in a spreadsheet reads as zero; the words do not."""
    client, _, _ = populated
    body = client.get("/export.csv?q=prairie").text

    data_line = [line for line in body.splitlines() if line.startswith("44")][0]
    assert "Not available" in data_line


def test_xlsx_export_is_a_readable_workbook(populated) -> None:
    client, _, _ = populated
    response = client.get("/export.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook["Prospects"]

    assert sheet["A1"].value == "Allstar Partners"
    assert sheet["A2"].value == "FQHC Prospect Intelligence"

    header_row = next(
        row for row in range(1, 12) if sheet.cell(row=row, column=1).value == "ICP score"
    )
    assert sheet.cell(row=header_row, column=2).value == "Organization"
    assert sheet.freeze_panes == f"A{header_row + 1}"

    names = [
        sheet.cell(row=row, column=2).value
        for row in range(header_row + 1, header_row + 6)
    ]
    assert "Erie Family Health" in names


def test_export_states_when_data_came_from_cache(populated) -> None:
    """An exported file outlives the banner, so it carries its own provenance."""
    client, db, _ = populated
    run = db.scalars(select(IngestRun)).one()
    run.used_cache = True
    run.source_reachable = False
    run.cache_date = datetime(2026, 5, 1, tzinfo=timezone.utc)
    db.commit()

    body = client.get("/export.csv").text
    assert "cached copy from 2026-05-01" in body


# ---------------------------------------------------------------------------
# Status and refresh
# ---------------------------------------------------------------------------


def test_data_status_reports_the_latest_run(populated) -> None:
    _, db, _ = populated
    status = data_status(db)

    assert status.has_data is True
    assert status.on_cached_data is False
    assert status.latest_run is not None


def test_cached_run_is_surfaced_as_a_banner(populated) -> None:
    client, db, _ = populated
    run = db.scalars(select(IngestRun)).one()
    run.used_cache = True
    run.source_reachable = False
    run.cache_date = datetime(2026, 5, 1, tzinfo=timezone.utc)
    run.message = "HRSA unreachable"
    db.commit()

    body = client.get("/").text
    assert "Showing cached data" in body
    assert "1 May 2026" in body


def test_refresh_status_endpoint_reports_idle(populated) -> None:
    client, _, _ = populated
    payload = client.get("/refresh/status").json()

    assert payload["running"] is False
    assert "percent" in payload


def test_healthz(populated) -> None:
    client, _, _ = populated
    assert client.get("/healthz").json()["status"] == "ok"


# ---------------------------------------------------------------------------
# What changed
# ---------------------------------------------------------------------------


def seed_change(session: Session, org_id: int, kind, summary: str, direction=None):
    from app.models import ChangeEvent

    session.add(
        ChangeEvent(
            organization_id=org_id,
            kind=kind,
            summary=summary,
            direction=direction,
        )
    )
    session.commit()


def test_changes_page_lists_events(populated) -> None:
    from app.models import ChangeKind

    client, db, erie = populated
    seed_change(db, erie.id, ChangeKind.SITES, "Opened 3 delivery sites (8 to 11)", 1)

    body = client.get("/changes").text

    assert "What changed" in body
    assert "Opened 3 delivery sites" in body
    assert "Delivery sites" in body
    assert "Erie Family Health" in body


def test_changes_page_filters_by_kind(populated) -> None:
    from app.models import ChangeKind

    client, db, erie = populated
    seed_change(db, erie.id, ChangeKind.SITES, "Opened 3 delivery sites (8 to 11)", 1)
    seed_change(db, erie.id, ChangeKind.AWARD, "Federal award increased 25%", 1)

    sites_only = client.get("/changes?kind=sites").text
    assert "Opened 3 delivery sites" in sites_only
    assert "Federal award increased" not in sites_only


def test_unknown_kind_falls_back_to_everything(populated) -> None:
    from app.models import ChangeKind

    client, db, erie = populated
    seed_change(db, erie.id, ChangeKind.SITES, "Opened 3 delivery sites (8 to 11)", 1)

    assert "Opened 3 delivery sites" in client.get("/changes?kind=nonsense").text


def test_changes_page_distinguishes_baseline_from_nothing_moving(populated) -> None:
    """"No events" reads very differently before and after a second run."""
    from app.models import IngestRun, RunStatus

    client, db, _ = populated
    db.add(
        IngestRun(
            stage="changes",
            status=RunStatus.SUCCESS,
            started_at=utcnow(),
            finished_at=utcnow(),
        )
    )
    db.commit()

    assert "Baseline recorded" in client.get("/changes").text

    db.add(
        IngestRun(
            stage="changes",
            status=RunStatus.SUCCESS,
            started_at=utcnow(),
            finished_at=utcnow(),
        )
    )
    db.commit()

    assert "Nothing has moved" in client.get("/changes").text


def test_navigation_shows_a_change_badge(populated) -> None:
    from app.models import ChangeKind

    client, db, erie = populated
    seed_change(db, erie.id, ChangeKind.SITES, "Opened 3 delivery sites", 1)

    body = client.get("/").text
    assert "nav__count--info" in body
    assert "What changed" in body


# ---------------------------------------------------------------------------
# Contacts export: one row per named person
# ---------------------------------------------------------------------------


def test_contacts_export_flattens_both_sources(populated) -> None:
    client, _, _ = populated
    body = client.get("/contacts.csv").text

    assert "MARIA T ALVAREZ" in body       # from a Form 990
    assert "Denise Whitaker" in body        # from a website
    assert "IRS Form 990 Part VII" in body
    assert "Organization website" in body


def test_every_contact_row_names_its_source(populated) -> None:
    """A filing and a web page are different claims, and the row that gets
    forwarded has to say which one it is."""
    client, _, _ = populated
    lines = client.get("/contacts.csv").text.splitlines()

    header = next(line for line in lines if line.startswith("Organization,State"))
    columns = header.split(",")
    assert "Source" in columns and "Source detail" in columns

    alvarez = next(line for line in lines if "MARIA T ALVAREZ" in line)
    whitaker = next(line for line in lines if "Denise Whitaker" in line)
    assert "Tax year 2023" in alvarez
    assert "https://example.org/board" in whitaker


def test_contacts_export_carries_the_caveat_into_the_file(populated) -> None:
    client, _, _ = populated
    body = client.get("/contacts.csv").text
    assert "should be confirmed before use" in body


def test_contacts_export_respects_the_current_filter(populated) -> None:
    client, _, _ = populated
    body = client.get("/contacts.csv?state=WI").text

    assert "Denise Whitaker" in body
    assert "MARIA T ALVAREZ" not in body
    assert "state in WI" in body


def test_contacts_export_marks_missing_emails(populated) -> None:
    """A 990 carries no email, and a blank cell would invite a guess."""
    client, _, _ = populated
    alvarez = next(
        line
        for line in client.get("/contacts.csv").text.splitlines()
        if "MARIA T ALVAREZ" in line
    )
    assert "Not available" in alvarez


def test_contacts_xlsx_is_a_readable_workbook(populated) -> None:
    from openpyxl import load_workbook

    client, _, _ = populated
    response = client.get("/contacts.xlsx")
    assert response.status_code == 200

    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    assert sheet.title == "Contacts"

    values = [
        cell.value
        for row in sheet.iter_rows(values_only=False)
        for cell in row
        if cell.value
    ]
    assert "MARIA T ALVAREZ" in values
    assert "Denise Whitaker" in values


def test_contacts_export_filename_is_distinct(populated) -> None:
    client, _, _ = populated
    disposition = client.get("/contacts.csv").headers["content-disposition"]
    assert "allstar-fqhc-contacts-" in disposition


def test_the_dashboard_offers_the_contacts_export(populated) -> None:
    client, _, _ = populated
    body = client.get("/").text
    assert 'href="/contacts.csv' in body and 'href="/contacts.xlsx' in body
