"""CSV and XLSX exports of the current filtered view.

Both formats open with a header block naming the company, the report, when it
was generated and which filters produced it, so a file that has been emailed
around still says what it is and how current it was.
"""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import Config
from app.formatting import NOT_AVAILABLE
from app.queries import ContactRow, DataStatus, Filters, ProspectRow
from pipeline.propublica import format_ein

COLUMNS: tuple[tuple[str, str], ...] = (
    ("ICP score", "score"),
    ("Organization", "name"),
    ("City", "city"),
    ("State", "state"),
    ("ZIP", "zip"),
    ("Delivery sites", "sites"),
    ("Patients", "patients"),
    ("Staff FTE", "staff_fte"),
    ("UDS year", "uds_year"),
    ("Grantee type", "grantee_type"),
    ("EIN", "ein"),
    ("EIN match status", "match_status"),
    ("EIN match confidence", "match_score"),
    ("Total revenue", "revenue"),
    ("Revenue tax year", "revenue_year"),
    ("Federal award", "award"),
    ("Phone", "phone"),
    ("Website", "website"),
)

# Brand colors, as openpyxl wants them (ARGB, no leading #).
BRAND_BLUE = "FF0094BB"
BRAND_GREEN = "FF6FC055"
BRAND_GRAY = "FF999999"


def _cell_values(row: ProspectRow) -> dict[str, object]:
    """Raw values for one row. Unknowns stay None until formatting."""
    organization = row.organization
    match = row.match
    return {
        "score": row.composite,
        "name": organization.name,
        "city": organization.city,
        "state": organization.state,
        "zip": organization.zip_code,
        "sites": organization.site_count,
        "patients": row.patients,
        "staff_fte": row.staff_fte,
        "uds_year": row.uds_year,
        "grantee_type": getattr(organization.grantee_type, "value", organization.grantee_type),
        "ein": format_ein(row.ein),
        "match_status": getattr(match.status, "value", match.status) if match else "unmatched",
        "match_score": match.score if match else None,
        "revenue": row.revenue,
        "revenue_year": row.revenue_year,
        "award": organization.federal_award_amount,
        "phone": organization.phone,
        "website": organization.website,
    }


def _header_block(
    config: Config,
    filters: Filters,
    row_count: int,
    status: DataStatus | None,
    generated_at: datetime | None = None,
) -> list[list[str]]:
    """The provenance block that opens every export."""
    stamp = (generated_at or datetime.now(timezone.utc)).strftime("%Y-%m-%d %H:%M UTC")
    block = [
        [config.app.company],
        [config.app.name],
        [f"Generated {stamp}"],
        [f"Filters: {filters.describe()}"],
        [f"Rows: {row_count:,}"],
    ]

    # If the underlying data came from cache, say so in the file itself -- an
    # exported spreadsheet outlives the banner on the screen.
    if status is not None:
        if status.on_cached_data and status.cache_date:
            block.append(
                [
                    "Source data: cached copy from "
                    f"{status.cache_date:%Y-%m-%d} (source was unreachable)"
                ]
            )
        elif status.latest_run and status.latest_run.finished_at:
            block.append(
                [f"Source data refreshed {status.latest_run.finished_at:%Y-%m-%d}"]
            )
    block.append([])
    return block


def to_csv(
    rows: list[ProspectRow],
    config: Config,
    filters: Filters,
    status: DataStatus | None = None,
    generated_at: datetime | None = None,
) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")

    for line in _header_block(config, filters, len(rows), status, generated_at):
        writer.writerow(line)

    writer.writerow([label for label, _ in COLUMNS])
    for row in rows:
        values = _cell_values(row)
        writer.writerow(
            [
                # Missing data is written as the same words the UI shows, so a
                # blank cell never gets read as a zero.
                NOT_AVAILABLE if values[key] is None else values[key]
                for _, key in COLUMNS
            ]
        )
    return buffer.getvalue()


MONEY_KEYS = frozenset({"revenue", "award", "compensation"})
DECIMAL_KEYS = frozenset({"score", "match_score"})

WIDTHS = {
    "name": 42,
    "organization": 42,
    "person": 28,
    "title": 30,
    "role": 26,
    "email": 30,
    "source": 24,
    "source_detail": 46,
    "city": 18,
    "website": 32,
    "match_status": 16,
    "grantee_type": 14,
    "revenue": 16,
    "award": 16,
    "patients": 12,
    "staff_fte": 12,
    "uds_year": 10,
    "compensation": 16,
    "phone": 16,
}


def _write_sheet(
    sheet,
    columns: tuple[tuple[str, str], ...],
    records: list[dict[str, object]],
    header_block: list[list[str]],
) -> None:
    """Write the provenance block, a branded header row and the data."""
    for index, line in enumerate(header_block, start=1):
        sheet.cell(row=index, column=1, value=line[0] if line else None)

    sheet.cell(row=1, column=1).font = Font(name="Open Sans", size=14, bold=True, color=BRAND_BLUE)
    sheet.cell(row=2, column=1).font = Font(name="Open Sans", size=11, bold=True)
    for index in range(3, len(header_block) + 1):
        sheet.cell(row=index, column=1).font = Font(name="Open Sans", size=9, color=BRAND_GRAY)

    header_row = len(header_block) + 1
    thin = Side(style="thin", color="FFE4E8EA")

    for column_index, (label, _) in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=label)
        cell.font = Font(name="Open Sans", size=10, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=BRAND_GREEN)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    for offset, values in enumerate(records, start=1):
        for column_index, (_, key) in enumerate(columns, start=1):
            value = values[key]
            cell = sheet.cell(
                row=header_row + offset,
                column=column_index,
                value=NOT_AVAILABLE if value is None else value,
            )
            cell.font = Font(name="Open Sans", size=10)
            if value is None:
                # Italic gray, so an unknown never reads as a reported zero.
                cell.font = Font(name="Open Sans", size=10, italic=True, color=BRAND_GRAY)
            elif key in MONEY_KEYS:
                cell.number_format = '"$"#,##0'
            elif key in DECIMAL_KEYS:
                cell.number_format = "0.0"

    for column_index, (label, key) in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = WIDTHS.get(
            key, max(len(label) + 2, 12)
        )

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)


def to_xlsx(
    rows: list[ProspectRow],
    config: Config,
    filters: Filters,
    status: DataStatus | None = None,
    generated_at: datetime | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Prospects"

    _write_sheet(
        sheet,
        COLUMNS,
        [_cell_values(row) for row in rows],
        _header_block(config, filters, len(rows), status, generated_at),
    )

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


# ---------------------------------------------------------------------------
# Contacts: one row per named person
# ---------------------------------------------------------------------------

CONTACT_COLUMNS: tuple[tuple[str, str], ...] = (
    ("Organization", "organization"),
    ("State", "state"),
    ("ICP score", "score"),
    ("Name", "person"),
    ("Title", "title"),
    ("Form 990 role", "role"),
    ("Email", "email"),
    ("Source", "source"),
    ("Source detail", "source_detail"),
    ("Reported compensation", "compensation"),
    ("Organization phone", "phone"),
    ("EIN", "ein"),
)


def _contact_values(contact: ContactRow) -> dict[str, object]:
    return {
        "organization": contact.organization.name,
        "state": contact.organization.state,
        "score": contact.composite,
        "person": contact.name,
        "title": contact.title,
        "role": contact.role,
        "email": contact.email,
        "source": contact.source,
        "source_detail": contact.source_detail,
        "compensation": contact.compensation,
        "phone": contact.organization.phone,
        "ein": format_ein(contact.ein),
    }


def contacts_to_csv(
    contacts: list[ContactRow],
    config: Config,
    filters: Filters,
    status: DataStatus | None = None,
    generated_at: datetime | None = None,
) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")

    for line in _header_block(config, filters, len(contacts), status, generated_at):
        writer.writerow(line)
    writer.writerow(_CONTACT_CAVEAT)
    writer.writerow([])

    writer.writerow([label for label, _ in CONTACT_COLUMNS])
    for contact in contacts:
        values = _contact_values(contact)
        writer.writerow(
            [
                NOT_AVAILABLE if values[key] is None else values[key]
                for _, key in CONTACT_COLUMNS
            ]
        )
    return buffer.getvalue()


def contacts_to_xlsx(
    contacts: list[ContactRow],
    config: Config,
    filters: Filters,
    status: DataStatus | None = None,
    generated_at: datetime | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contacts"

    header_block = _header_block(config, filters, len(contacts), status, generated_at)
    # Slot the caveat in before the closing blank line of the block.
    header_block = [*header_block[:-1], list(_CONTACT_CAVEAT), []]

    _write_sheet(
        sheet, CONTACT_COLUMNS, [_contact_values(c) for c in contacts], header_block
    )

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


# Travels with the file, because a contact list gets forwarded and the screen
# that explained it does not.
_CONTACT_CAVEAT = [
    "Source column: Form 990 rows come from a signed federal filing and "
    "describe the tax year shown. Website rows were read from the page linked "
    "in Source detail and should be confirmed before use. Emails appear only "
    "where the organization published them itself."
]


def export_filename(extension: str, generated_at: datetime | None = None) -> str:
    stamp = (generated_at or datetime.now(timezone.utc)).strftime("%Y-%m-%d")
    return f"allstar-fqhc-prospects-{stamp}.{extension}"


def contacts_filename(extension: str, generated_at: datetime | None = None) -> str:
    stamp = (generated_at or datetime.now(timezone.utc)).strftime("%Y-%m-%d")
    return f"allstar-fqhc-contacts-{stamp}.{extension}"
